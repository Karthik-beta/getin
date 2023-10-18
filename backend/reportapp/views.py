from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Count, Q
from rest_framework import generics
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from django.utils.dateparse import parse_date
from datetime import datetime, timedelta
import openpyxl
from django.views.generic import View
from django.http import HttpResponse
from rest_framework import status
from dateutil.relativedelta import relativedelta




from .models import Logs, EmployeeData, AttendanceRecord
from .serializers import LogsSerializer, EmployeeDataSerializer, AttendanceRecordSerializer



class AttendanceRecordPagination(PageNumberPagination):
    page_size = 4  # Adjust the number of records per page as needed
    page_size_query_param = 'page_size'
    max_page_size = 100  # Set the maximum page size if desired

class AttendanceRecordListView(generics.ListAPIView):
    queryset = AttendanceRecord.objects.order_by('-logdate') 
    serializer_class = AttendanceRecordSerializer
    pagination_class = AttendanceRecordPagination
    filter_backends = [DjangoFilterBackend]  
    filterset_fields = ['employeeid', 'employee_name', 'logdate', 'company', 'logdate', 'shift_status']  

    def parse_date_with_format(self, date_string, format):
        try:
            return datetime.strptime(date_string, format).date()
        except ValueError:
            return None

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get('search', None)
        date_query = self.request.query_params.get('date', None)  

        if date_query:
            try:
                parsed_date = self.parse_date_with_format(date_query, "%Y/%m/%d")  
                next_day = parsed_date + timedelta(days=1)  

                queryset = queryset.filter(
                    logdate__gte=parsed_date,
                    logdate__lt=next_day
                )
            except Exception as e:
                print("Error parsing date:", e)

        if search_query:
            queryset = queryset.filter(
                Q(employee_name__icontains=search_query) |
                Q(employeeid__icontains=search_query)
            )
        return queryset


class AutocompleteView(APIView):  
    def get(self, request):
        search_term = request.query_params.get('term')
        if search_term:
            attendance_suggestions = AttendanceRecord.objects.filter(
                Q(employee_name__icontains=search_term) |
                Q(employeeid__icontains=search_term)
            )[:5]  # Limit the number of attendance suggestions to 5
            attendance_suggestions = list(attendance_suggestions)

            suggestions = []

            for item in attendance_suggestions:
                value = f"{item.employee_name} - {item.employeeid}"
                suggestions.append({'id': item.id, 'value': value})

            return Response(suggestions)
        return Response([])









class AttendanceSummaryView(APIView):
    def get(self, request):
        date = request.query_params.get('date')
        if not date:
            return Response({"error": "Please provide a 'date' parameter."}, status=400)

        # Calculate attendance summary
        summary = AttendanceRecord.objects.filter(logdate=date).aggregate(
            total_present=Count('id', filter=Q(shift_status='P') | (Q(direction='in') & Q(logdate=date))),
            total_absent=Count('id', filter=Q(direction='out') | (Q(direction='in') & ~Q(logdate=date))),
            total_late_entry=Count('id', filter=(Q(late_entry__gt='00:00:00') | Q(early_exit__gt='00:00:00')) & (Q(shift_status='P') | (Q(direction='in') & Q(logdate=date)))),
            total_early_exit=Count('id', filter=(Q(late_entry__gt='00:00:00') | Q(early_exit__gt='00:00:00')) & (Q(shift_status='P') | (Q(direction='in') & Q(logdate=date)))),
        )

        return Response(summary)




class EmployeeDataPagination(PageNumberPagination):
    page_size = 10  # Adjust the number of records per page as needed
    page_size_query_param = 'page_size'
    max_page_size = 100  # Set the maximum page size if desired


class EmployeeDataListCreateView(generics.ListCreateAPIView):
    queryset = EmployeeData.objects.order_by('-employee_id')
    serializer_class = EmployeeDataSerializer
    pagination_class = EmployeeDataPagination 
    filter_backends = [DjangoFilterBackend]  # Add this line to include the filter backend
    filterset_fields = ['employee_id', 'employee_name', 'location', 'company']

class EmployeeDataRetrieveUpdateDeleteView(generics.RetrieveAPIView, generics.UpdateAPIView, generics.DestroyAPIView):
    queryset = EmployeeData.objects.all()
    serializer_class = EmployeeDataSerializer
    lookup_field = 'employee_id'  # Use 'employee_id' as the lookup field

    def options(self, request, *args, **kwargs):
        response = Response()
        # Add CORS headers to the response
        response["Access-Control-Allow-Origin"] = "http://localhost:4200"  # Replace with your frontend's domain
        response["Access-Control-Allow-Methods"] = "GET, PUT, DELETE, OPTIONS"
        response["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
        return response
    




class ExportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        headers = ["Employee ID", "Employee Name", "Log Date", "Company", "Shift Status"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.employeeid)
            ws.cell(row=row_num, column=2, value=record.employee_name)
            ws.cell(row=row_num, column=3, value=record.logdate)
            ws.cell(row=row_num, column=4, value=record.company)
            ws.cell(row=row_num, column=5, value=record.shift_status)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=attendance_records.xlsx"
        wb.save(response)

        return response

    def get_filtered_queryset(self, request):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get('search', None)
        date_query = self.request.query_params.get('date', None)  

        if date_query:
            try:
                parsed_date = self.parse_date_with_format(date_query, "%Y/%m/%d")  
                next_day = parsed_date + timedelta(days=1)  

                queryset = queryset.filter(
                    logdate__gte=parsed_date,
                    logdate__lt=next_day
                )
            except Exception as e:
                print("Error parsing date:", e)

        if search_query:
            queryset = queryset.filter(
                Q(employee_name__icontains=search_query) |
                Q(employeeid__icontains=search_query)
            )
        return queryset





class AttendanceStatisticsAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Calculate the latest date in the AttendanceRecord model
        latest_date = AttendanceRecord.objects.latest('logdate').logdate

        # Query to calculate statistics
        attendance_stats = AttendanceRecord.objects.filter(logdate=latest_date).aggregate(
            total_employees=Count('id'),
            present_count=Count('id', filter=~Q(first_logtime__isnull=True)),
            absent_count=Count('id', filter=Q(first_logtime__isnull=True)),
            late_entry_count=Count('id', filter=~Q(late_entry="00:00")),
            early_exit_count=Count('id', filter=~Q(early_exit="00:00")),
            overtime_count=Count('id', filter=~Q(overtime="00:00"))
        )

        # Extract the counts
        total_employees = attendance_stats['total_employees']
        present_count = attendance_stats['present_count']
        absent_count = attendance_stats['absent_count']
        late_entry_count = attendance_stats['late_entry_count']
        early_exit_count = attendance_stats['early_exit_count']
        overtime_count = attendance_stats['overtime_count']

        # Prepare the response data
        response_data = {
            'latest_date': latest_date,
            'total_employees': total_employees,
            'early_exit': early_exit_count,
            'present': present_count,
            'late_entry': late_entry_count,
            'absent': absent_count,
            'overtime': overtime_count
        }

        return Response(response_data)





class AttendanceStatisticsMonthAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Calculate the date 30 days ago from today
        thirty_days_ago = datetime.now() - timedelta(days=66)

        # Query to calculate statistics for each date in the last 30 days
        attendance_stats = AttendanceRecord.objects.filter(logdate__gte=thirty_days_ago).values('logdate').annotate(
            total_employees=Count('id'),
            present_count=Count('id', filter=~Q(first_logtime__isnull=True)),
            absent_count=Count('id', filter=Q(first_logtime__isnull=True)),
            late_entry_count=Count('id', filter=~Q(late_entry="00:00")),
            early_exit_count=Count('id', filter=~Q(early_exit="00:00")),
            overtime_count=Count('id', filter=~Q(overtime="00:00"))
        )

        # Prepare the response data as a list of dictionaries
        response_data = []
        for stat in attendance_stats:
            response_data.append({
                'logdate': stat['logdate'].strftime('%Y-%m-%d'),
                'total_employees_count': stat['total_employees'],
                'present_count': stat['present_count'],
                'absent_count': stat['absent_count'],
                'late_entry_count': stat['late_entry_count'],
                'early_exit_count': stat['early_exit_count'],
                'overtime_count': stat['overtime_count']
            })

        return Response(response_data)
    


    
    



class DownloadEmployeeData(APIView):
    def get(self, request):
        # Query the EmployeeData model to retrieve the data
        employee_data = EmployeeData.objects.all()

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add column headers to the worksheet
        ws.append(['ID', 'Employee ID', 'Device Enroll ID', 'Employee Name', 'Company', 'Location'])

        # Add data from the EmployeeData model to the worksheet
        for data in employee_data:
            ws.append([data.id, data.employee_id, data.device_enroll_id, data.employee_name, data.company, data.location])

        # Save the workbook to a response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=employee_data.xlsx'
        wb.save(response)

        return response
    


class DownloadAttendanceData(APIView):
    def get(self, request):
        # Query the AttendanceRecord model to retrieve the data
        attendance_data = AttendanceRecord.objects.all()

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add column headers to the worksheet
        ws.append(['Employee ID', 'Employee Name', 'Log Date', 'first_logtime', 'last_logtime', 'Direction', 'Shift Status', 'Late Entry', 'Early Exit', 'Overtime', 'Company', 'location', 'shift_status'])

        # Add data from the AttendanceRecord model to the worksheet
        for data in attendance_data:
            ws.append([data.employeeid, data.employee_name, data.logdate, data.first_logtime, data.last_logtime, data.direction, data.shift_status, data.late_entry, data.early_exit, data.overtime, data.company, data.location, data.shift_status])

        # Save the workbook to a response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=attendance_data.xlsx'
        wb.save(response)

        return response
    









class GenderCountAPI(APIView):
    def get(self, request):
        # Calculate gender counts
        male_count = EmployeeData.objects.filter(gender='MALE').count()
        female_count = EmployeeData.objects.filter(gender='FEMALE').count()
        others_count = EmployeeData.objects.exclude(gender__in=['MALE', 'FEMALE']).count()

        # Calculate onboarding, active, resigned, and absconded counts based on the status column
        total_count = EmployeeData.objects.count()
        active_count = EmployeeData.objects.filter(status='ACTIVE').count()
        resigned_count = EmployeeData.objects.filter(status='RESIGNED').count()
        absconded_count = EmployeeData.objects.filter(status='ABSCONDED').count()

        # Create a dictionary with all the counts
        gender_counts = {
            'male': male_count,
            'female': female_count,
            'others': others_count,
            'total': total_count,
            'active': active_count,
            'resigned': resigned_count,
            'absconded': absconded_count,
        }

        return Response(gender_counts, status=status.HTTP_200_OK)